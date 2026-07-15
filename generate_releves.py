"""
Genere un releve de notes (.xlsx) par etudiant a partir d'un ou deux PV de
deliberation (Semestre 1 et/ou Semestre 2), pour n'importe quelle filiere
(le nombre d'UE/ECUE par semestre est detecte dynamiquement depuis le PV,
le bloc UE du gabarit est redimensionne en consequence).

Usage CLI:
    python generate_releves.py <pv_s1.xlsx> [pv_s2.xlsx] [--template gabarit.xlsx] [--out dossier]

Ce module est aussi importe par app.py (interface web).

Hypotheses actuelles (a ajuster si la situation change) :
- "Decision du jury" / mention laissees vides (remplissage manuel par le jury).
- Nom & Prenom recopie tel quel depuis le PV (ordre "Nom Prenom", pas inverse).
- Une UE est consideree "Validee/Capitalisee" (V/C) si ses Credits Capitalises
  (colonne du PV) sont > 0 ; a corriger si la regle exacte de l'etablissement
  differe.
- Si un etudiant n'apparait que dans un des deux PV fournis, le releve est
  quand meme genere avec les donnees disponibles (l'autre semestre reste vide)
  et un avertissement est renvoye par generate_all().
- Date/lieu de naissance et N CIN/Passeport viennent du fichier de
  coordonnees des etudiants (optionnel) si fourni, en faisant correspondre
  les etudiants par nom ; sinon laisses vides (remplissage manuel).
"""
import copy
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent
PV_SHEET_NAME = "PV_Deliberation"
NAME_COLUMN = 2  # colonne B : "Nom & Prenom"

# Bloc UE/ECUE du gabarit : ligne de depart, et ligne ou commence le pied de
# page (TOTAL/Moyenne Annuelle/Decision...) dans le gabarit d'origine. Le
# bloc est redimensionne dynamiquement (insert_rows/delete_rows) pour
# s'adapter au nombre reel d'UE/ECUE trouve dans le PV, donc ces constantes
# ne sont que le point de depart du calcul, pas une contrainte sur le PV.
UE_BLOCK_FIRST_ROW = 19
FOOTER_FIRST_ROW = 40  # ligne "TOTAL" dans le gabarit d'origine

# Filieres proposees dans l'interface -> code utilise pour retrouver l'onglet
# correspondant dans le fichier de coordonnees des etudiants (ex. onglet
# "L1-BD" pour le code "BD" au niveau L1).
FILIERES = {
    "Licence Nationale en Génie Logiciel et Système d'Information": "GLSI",
    "Licence Nationale en Big Data et Analyse des Données": "BD",
    "Licence en Ingénierie des Systèmes et Réseaux": "RSYS",
    "Mastère en Cloud Computing et Virtualisation": "CLOUD",
    "Mastère en Cybersécurité": "CYBER",
}

# Niveau -> (chiffre, suffixe ordinal) pour l'Attestation de Reussite
# (ex. L1 -> "1" + "ère" = "1ère année").
NIVEAU_TEXTE = {
    "L1": ("1", "ère"), "L2": ("2", "ème"), "L3": ("3", "ème"),
    "M1": ("1", "ère"), "M2": ("2", "ème"),
}

# Annees terminales de cycle : c'est a l'issue de ces annees qu'un diplome
# national est delivre (3e annee de Licence, 2e annee de Mastere).
FINAL_YEARS = {"L3", "M2"}

# Niveaux proposes dans l'interface selon le cycle (3 ans de Licence, 2 ans
# de Mastere).
NIVEAUX_PAR_CYCLE = {
    "Licence": [("L1", "1ère année"), ("L2", "2ème année"), ("L3", "3ème année")],
    "Mastère": [("M1", "1ère année"), ("M2", "2ème année")],
}


def cycle_de_filiere(label):
    return "Licence" if label.startswith("Licence") else "Mastère" if label.startswith("Mastère") else ""


@dataclass
class Ecue:
    code: str
    name: str
    coef: float
    credit: float
    cc_col: int
    ex_col: int
    moy_col: int
    cr_col: int


@dataclass
class UniteEnseignement:
    label: str
    semestre: str
    coef: float
    credit: float
    ecues: list
    moyenne_ue_col: int
    credits_cap_col: int


@dataclass
class StudentRecord:
    num: object
    nom_prenom: str
    row: int
    grades: dict          # ecue_code -> (cc, ex, moy, cr)
    ue_results: dict       # ue_label -> (moyenne_ue, credits_capitalises)
    moyenne_semestre: float
    total_credits_semestre: int


def _clean(text):
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def normalize_name(text):
    """Normalise un nom pour comparaison (insensible aux accents, a la
    casse et aux espaces multiples)."""
    text = unicodedata.normalize("NFKD", _clean(text))
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text).strip().upper()


def parse_label_value(ws, rows, label):
    """Cherche 'label :' dans les lignes donnees, meme quand label et valeur
    sont dans la meme cellule ou dans deux cellules separees de la meme ligne."""
    for row in rows:
        cells = [ws.cell(row=row, column=c) for c in range(1, ws.max_column + 1)]
        for idx, cell in enumerate(cells):
            val = cell.value
            if not isinstance(val, str) or label not in val:
                continue
            after = val.split(label, 1)[1].strip(" : ")
            if after:
                return _clean(after)
            for nxt in cells[idx + 1:]:
                if nxt.value not in (None, ""):
                    return _clean(nxt.value)
    return ""


def merges_starting_at(ws, row, min_col=1, max_col=None):
    max_col = max_col or ws.max_column
    spans = [mc for mc in ws.merged_cells.ranges
             if mc.min_row == row and min_col <= mc.min_col <= max_col]
    return sorted(spans, key=lambda r: r.min_col)


HEADER_SCAN_ROWS = 30  # profondeur de balayage de la zone d'en-tete du PV
ECUE_WIDTH = 4         # chaque ECUE occupe 4 colonnes (CC, EX, Moy, Cr)

# Intitules de synthese (colonnes de recap) : servent de garde d'arret quand
# on reconstruit les ECUE d'une UE, pour ne jamais confondre une ECUE avec une
# colonne de recapitulatif (utile quand ces colonnes ne sont pas fusionnees).
_SUMMARY_HEADERS = ("Moyenne UE", "Moyenne Semestre", "Moyenne Annuelle",
                    "Total", "Observation")


def _is_summary_header(val):
    v = val.replace("é", "e").replace("è", "e")  # 'Credits'/'Credits Capitalises'
    return v.startswith("Credits Capitalis") or any(val.startswith(h) for h in _SUMMARY_HEADERS)


def _ue_markers(ws):
    """Chaque UE se termine par une colonne 'Moyenne UE' (suivie de 'Credits
    Capitalises') sur la ligne des ECUE : on repere les UE par ces marqueurs.

    On balaie les VALEURS des cellules (et pas seulement les cellules
    fusionnees) : certains PV laissent un semestre entier NON fusionne (vu sur
    un PV L2 ou tout le Semestre 2 etait en cellules simples), ce qui faisait
    disparaitre ce semestre. Un 'Moyenne UE' fusionne n'a sa valeur que sur sa
    cellule d'ancrage, donc ce balayage ne cree pas de doublon.

    Marche aussi quand les UE d'un semestre sont sur une LIGNE differente de
    celles d'un autre (PV a semestres cote a cote, decales d'une ligne).
    Renvoie [(moyenne_ue_col, ecue_row), ...] trie par colonne (ecue_row =
    ligne des ECUE de cette UE)."""
    markers = []
    for row in range(1, HEADER_SCAN_ROWS + 1):
        for col in range(1, ws.max_column + 1):
            if _clean(ws.cell(row=row, column=col).value) == "Moyenne UE":
                markers.append((col, row))
    markers.sort()
    return markers


def _semester_boundaries(ws, credits_cap_cols):
    """Colonnes 'Moyenne Semestre N' qui cloturent un bloc d'UE (donc
    precedees d'une colonne 'Credits Capitalises' d'UE) : ce sont les vrais
    recapitulatifs de semestre, par opposition au bloc de synthese annuelle
    place tout a droite (Moyenne Annuelle, Decisions...) qui reprend les memes
    intitules. Renvoie [(col, N), ...] trie par colonne.

    Insensible aux fusions (meme balayage par valeur que _ue_markers) : le
    recap 'Moyenne Semestre N' peut lui aussi etre non fusionne."""
    bounds = []
    seen = set()
    for row in range(1, HEADER_SCAN_ROWS + 1):
        for col in range(1, ws.max_column + 1):
            m = re.match(r"Moyenne\s+Semestre\s+(\d)",
                         _clean(ws.cell(row=row, column=col).value))
            if m and (col - 1) in credits_cap_cols and col not in seen:
                seen.add(col)
                bounds.append((col, int(m.group(1))))
    bounds.sort()
    return bounds


def _build_ues(ws):
    """Detecte toutes les UE/ECUE du PV a partir des marqueurs 'Moyenne UE'.
    Renvoie une liste de tuples (UniteEnseignement, ecue_row) triee de gauche
    a droite (= Semestre 1 puis Semestre 2), ou None si aucun marqueur (le PV
    ne suit pas ce gabarit -> on retombe sur l'ancien parseur)."""
    markers = _ue_markers(ws)
    if not markers:
        return None
    credits_cap_cols = {mue_col + 1 for mue_col, _ in markers}
    boundaries = _semester_boundaries(ws, credits_cap_cols)

    def semester_of(mue_col):
        for col, n in boundaries:
            if mue_col <= col:
                return n
        return boundaries[-1][1] if boundaries else None

    built = []
    prev_cap = 0
    for mue_col, ecue_row in markers:
        ue_name_row = ecue_row - 1
        # ECUE de cette UE : chaque ECUE occupe ECUE_WIDTH colonnes et la
        # derniere se termine juste avant 'Moyenne UE'. On remonte donc de
        # ECUE_WIDTH en ECUE_WIDTH depuis mue_col jusqu'a la fin de l'UE
        # precedente (prev_cap). Approche geometrique (et non basee sur les
        # fusions) : marche meme quand les ECUE d'un semestre ne sont pas
        # fusionnees. On s'arrete a une cellule vide ou a un intitule de
        # synthese (recap de semestre non fusionne intercale entre deux UE).
        ecue_cols = []
        col = mue_col - ECUE_WIDTH
        while col > prev_cap:
            val = _clean(ws.cell(row=ecue_row, column=col).value)
            if not val or _is_summary_header(val):
                break
            ecue_cols.append(col)
            col -= ECUE_WIDTH
        ecue_cols.reverse()
        prev_cap = mue_col + 1
        if not ecue_cols:
            continue
        ecues = []
        for cc in ecue_cols:
            raw = _clean(ws.cell(row=ecue_row, column=cc).value).replace("\n", " ")
            m = re.search(r"\(([^)]+)\)\s*$", raw)
            code = m.group(1) if m else raw
            name = raw.split("(")[0].strip()
            ecues.append(Ecue(
                code=code,
                name=name,
                coef=ws.cell(row=ecue_row + 1, column=cc).value,
                credit=ws.cell(row=ecue_row + 2, column=cc).value,
                cc_col=cc,
                ex_col=cc + 1,
                moy_col=cc + 2,
                cr_col=cc + 3,
            ))
        # Nom / coef / credit de l'UE : sur la ligne du nom d'UE, a la colonne
        # de depart de sa 1ere ECUE. Si le nom d'UE est fusionne (cas courant),
        # on prend l'ancre de la fusion ; sinon la colonne de la 1ere ECUE.
        first_col = ecue_cols[0]
        name_mc = next((mc for mc in ws.merged_cells.ranges
                        if mc.min_row == ue_name_row and mc.min_col <= first_col <= mc.max_col), None)
        anchor = name_mc.min_col if name_mc else first_col
        n = semester_of(mue_col)
        built.append((UniteEnseignement(
            label=_clean(ws.cell(row=ue_name_row, column=anchor).value),
            semestre=f"Semestre {n}" if n else "",
            coef=ws.cell(row=ecue_row + 3, column=anchor).value,
            credit=ws.cell(row=ecue_row + 4, column=anchor).value,
            ecues=ecues,
            moyenne_ue_col=mue_col,
            credits_cap_col=mue_col + 1,
        ), ecue_row))
    return built


def _block_name_col(ws, header_row, min_ue_col):
    """Colonne 'Nom & Prenom' d'un bloc (sous-tableau d'un semestre) = l'en-tete
    'Nom' de header_row la plus proche a gauche de la 1ere colonne d'UE du bloc.
    Un PV a semestres cote a cote a une colonne de noms par bloc."""
    candidates = [c for c in range(1, ws.max_column + 1)
                  if "Nom" in str(ws.cell(row=header_row, column=c).value or "") and c < min_ue_col]
    return max(candidates) if candidates else None


def parse_pv(path, sheet_name=PV_SHEET_NAME):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

    meta = {
        "domaine": parse_label_value(ws, range(1, 10), "Domaine"),
        "filiere": parse_label_value(ws, range(1, 10), "Filière"),
        "specialite": parse_label_value(ws, range(1, 10), "Spécialité"),
    }

    built = _build_ues(ws)
    if built is None:
        return _parse_pv_legacy(ws, meta)

    built.sort(key=lambda t: t[0].ecues[0].cc_col)  # gauche -> droite
    ues = [ue for ue, _ in built]

    # Blocs = sous-tableaux physiques (chacun sa colonne de noms et ses lignes
    # etudiants), regroupes par ligne des ECUE. Le semestre 2 peut etre un
    # bloc distinct, decale d'une ligne : on lit chaque bloc a sa propre place
    # puis on fusionne les etudiants par nom (comme la fusion de 2 PV separes).
    blocks = {}
    for ue, ecue_row in built:
        blocks.setdefault(ecue_row, []).append(ue)

    merged = {}   # nom_normalise -> {display, num, grades, ue_results}
    order = []
    for ecue_row in sorted(blocks):
        block_ues = blocks[ecue_row]
        header_row = ecue_row + 5   # meme geometrie interne que le gabarit d'origine
        name_col = _block_name_col(ws, header_row, min(u.ecues[0].cc_col for u in block_ues))
        if name_col is None:
            continue
        row = header_row + 1
        while True:
            nom = ws.cell(row=row, column=name_col).value
            if nom is None or _clean(nom) == "":
                break
            key = normalize_name(nom)
            rec = merged.get(key)
            if rec is None:
                rec = {"display": _clean(nom),
                       "num": ws.cell(row=row, column=name_col - 1).value,
                       "grades": {}, "ue_results": {}}
                merged[key] = rec
                order.append(key)
            for ue in block_ues:
                for ec in ue.ecues:
                    rec["grades"][ec.code] = (
                        ws.cell(row=row, column=ec.cc_col).value or 0,
                        ws.cell(row=row, column=ec.ex_col).value or 0,
                        ws.cell(row=row, column=ec.moy_col).value or 0,
                        ws.cell(row=row, column=ec.cr_col).value or 0,
                    )
                rec["ue_results"][ue.label] = (
                    ws.cell(row=row, column=ue.moyenne_ue_col).value or 0,
                    ws.cell(row=row, column=ue.credits_cap_col).value or 0,
                )
            row += 1

    students = [StudentRecord(
        num=merged[k]["num"], nom_prenom=merged[k]["display"], row=0,
        grades=merged[k]["grades"], ue_results=merged[k]["ue_results"],
        moyenne_semestre=0, total_credits_semestre=0,
    ) for k in order]

    return ues, students, meta


def _parse_pv_legacy(ws, meta):
    """Ancien parseur (un seul tableau, une colonne de noms, tous les
    semestres sur la meme ligne d'en-tete UE et la meme ligne etudiant).
    Utilise en repli si le PV n'expose pas de colonnes 'Moyenne UE'."""
    header_row = None
    for row in range(1, 30):
        val = ws.cell(row=row, column=NAME_COLUMN).value
        if val and "Nom" in str(val):
            header_row = row
            break
    if header_row is None:
        raise ValueError("Ligne d'en-tete introuvable (colonne B contenant 'Nom').")

    ue_name_row = header_row - 6
    ecue_name_row = header_row - 5
    coef_ecue_row = header_row - 4
    credit_ecue_row = header_row - 3
    coef_ue_row = header_row - 2
    credit_ue_row = header_row - 1
    semester_row = ue_name_row - 1
    summary_max_row = header_row

    ues = []
    for ue_mc in merges_starting_at(ws, ue_name_row, min_col=NAME_COLUMN + 1):
        if ue_mc.max_row != ue_name_row:
            continue  # exclut les merges verticaux de synthese (Moyenne Semestre, Total Credits)
        label = _clean(ws.cell(row=ue_name_row, column=ue_mc.min_col).value)
        if not label:
            continue

        # Le merge "Semestre X" ne couvre pas toujours toutes les colonnes
        # des UE de ce semestre (vu sur certains PV) -> on rattache l'UE au
        # dernier "Semestre X" dont le merge demarre avant ou a sa colonne,
        # plutot que d'exiger une inclusion stricte dans le merge.
        semestre = ""
        for sem_mc in sorted(merges_starting_at(ws, semester_row), key=lambda mc: mc.min_col):
            if sem_mc.min_col > ue_mc.min_col:
                break
            semestre = _clean(ws.cell(row=semester_row, column=sem_mc.min_col).value)

        ecue_spans = [
            mc for mc in merges_starting_at(ws, ecue_name_row, ue_mc.min_col, ue_mc.max_col)
            if mc.max_row == ecue_name_row
        ]
        ecues = []
        for span in ecue_spans:
            raw = _clean(ws.cell(row=ecue_name_row, column=span.min_col).value).replace("\n", " ")
            m = re.search(r"\(([^)]+)\)\s*$", raw)
            code = m.group(1) if m else raw
            name = raw.split("(")[0].strip()
            ecues.append(Ecue(
                code=code,
                name=name,
                coef=ws.cell(row=coef_ecue_row, column=span.min_col).value,
                credit=ws.cell(row=credit_ecue_row, column=span.min_col).value,
                cc_col=span.min_col,
                ex_col=span.min_col + 1,
                moy_col=span.min_col + 2,
                cr_col=span.min_col + 3,
            ))

        # Les colonnes "Moyenne UE" / "Credits Capitalises" suivent toujours
        # directement la derniere ECUE. On ne se fie pas a la largeur du merge
        # du nom d'UE (row11) car elle est incoherente d'un bloc a l'autre
        # dans ce fichier (parfois elle inclut ces 2 colonnes, parfois non).
        last_ecue_max_col = max(span.max_col for span in ecue_spans)
        moyenne_ue_col, credits_cap_col = last_ecue_max_col + 1, last_ecue_max_col + 2
        found = {
            mc.min_col for mc in ws.merged_cells.ranges
            if mc.min_row in (ue_name_row, ecue_name_row) and mc.max_row == summary_max_row
            and mc.min_col in (moyenne_ue_col, credits_cap_col)
        }
        if found != {moyenne_ue_col, credits_cap_col}:
            raise ValueError(
                f"UE '{label}': colonnes de synthese attendues en {moyenne_ue_col}/{credits_cap_col} "
                f"non confirmees par un merge (Moyenne UE / Credits Capitalises)."
            )

        ues.append(UniteEnseignement(
            label=label,
            semestre=semestre,
            coef=ws.cell(row=coef_ue_row, column=ue_mc.min_col).value,
            credit=ws.cell(row=credit_ue_row, column=ue_mc.min_col).value,
            ecues=ecues,
            moyenne_ue_col=moyenne_ue_col,
            credits_cap_col=credits_cap_col,
        ))

    semestre_summary = []
    for mc in ws.merged_cells.ranges:
        if mc.max_row == summary_max_row and mc.min_row in (ue_name_row, ecue_name_row):
            txt = _clean(ws.cell(row=mc.min_row, column=mc.min_col).value)
            if "Moyenne Semestre" in txt or "Total des Cr" in txt:
                semestre_summary.append((txt, mc.min_col))
    moyenne_sem_col = next(c for t, c in semestre_summary if "Moyenne" in t)
    total_credits_col = next(c for t, c in semestre_summary if "Total" in t)

    students = []
    row = header_row + 1
    while True:
        nom = ws.cell(row=row, column=NAME_COLUMN).value
        if nom is None or _clean(nom) == "":
            break
        grades, ue_results = {}, {}
        for ue in ues:
            for ec in ue.ecues:
                grades[ec.code] = (
                    ws.cell(row=row, column=ec.cc_col).value or 0,
                    ws.cell(row=row, column=ec.ex_col).value or 0,
                    ws.cell(row=row, column=ec.moy_col).value or 0,
                    ws.cell(row=row, column=ec.cr_col).value or 0,
                )
            ue_results[ue.label] = (
                ws.cell(row=row, column=ue.moyenne_ue_col).value or 0,
                ws.cell(row=row, column=ue.credits_cap_col).value or 0,
            )
        students.append(StudentRecord(
            num=ws.cell(row=row, column=1).value,
            nom_prenom=_clean(nom),
            row=row,
            grades=grades,
            ue_results=ue_results,
            moyenne_semestre=ws.cell(row=row, column=moyenne_sem_col).value or 0,
            total_credits_semestre=ws.cell(row=row, column=total_credits_col).value or 0,
        ))
        row += 1

    return ues, students, meta


def parse_filename_meta(path):
    name = path.stem
    annee = ""
    m = re.search(r"AU(\d{2})-(\d{2})", name)
    if m:
        annee = f"20{m.group(1)}-20{m.group(2)}"
    niveau = ""
    m = re.search(r"\b([LM][1-3])\b", name.upper())
    if m:
        niveau = m.group(1)
    return annee, niveau


def _academic_year_start(today=None):
    """Annee de debut de l'annee universitaire en cours : une AU commence en
    septembre (mois >= 8), donc de janvier a juillet on est encore dans l'AU
    ouverte l'annee civile precedente."""
    today = today or date.today()
    return today.year if today.month >= 8 else today.year - 1


def current_academic_year(today=None):
    """Annee universitaire en cours au format 'AAAA-AAAA' (ex. '2025-2026')."""
    start = _academic_year_start(today)
    return f"{start}-{start + 1}"


def academic_year_choices(today=None, back=2, forward=1):
    """Liste d'annees universitaires proposees dans l'interface, centree sur
    l'annee en cours (quelques annees en arriere + une en avant), la plus
    recente d'abord."""
    start = _academic_year_start(today)
    return [f"{y}-{y + 1}" for y in range(start + forward, start - back - 1, -1)]


def sanitize_filename(text):
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")


def sanitize_sheet_name(text):
    return re.sub(r"[\\/*?:\[\]]", "", text)[:31] or "Etudiant"


def to_number(value):
    """Convertit une valeur de cellule en float, en tolerant les nombres
    saisis comme texte avec virgule decimale (ex. '16,5') vus sur certains PV."""
    if value in (None, ""):
        return 0
    if isinstance(value, str):
        value = value.strip().replace(",", ".")
    return float(value)


def format_note(value):
    return round(to_number(value), 2) if value not in (None, "") else 0


def semester_digit(label):
    m = re.search(r"(\d)", label or "")
    return m.group(1) if m else None


def merge_students(students_per_pv):
    """Fusionne les etudiants de plusieurs PV (par Nom & Prenom). Renvoie un
    dict nom_prenom -> {"grades", "ue_results", "present_in"}."""
    merged = {}
    for idx, students in enumerate(students_per_pv):
        for s in students:
            rec = merged.setdefault(s.nom_prenom, {"grades": {}, "ue_results": {}, "present_in": set()})
            rec["grades"].update(s.grades)
            rec["ue_results"].update(s.ue_results)
            rec["present_in"].add(idx)
    return merged


def annual_aggregates(ue_results, ue_by_label):
    """Moyenne ponderee (par coefficient d'UE) et total des credits
    capitalises, sur toutes les UE presentes dans ue_results."""
    total_weight = 0.0
    weighted_sum = 0.0
    total_credits = 0
    for label, (moyenne_ue, credits_cap) in ue_results.items():
        ue = ue_by_label[label]
        coef = to_number(ue.coef)
        total_weight += coef
        weighted_sum += to_number(moyenne_ue) * coef
        total_credits += int(to_number(credits_cap))
    moyenne = round(weighted_sum / total_weight, 2) if total_weight else 0
    return moyenne, total_credits


def mention_from_average(moyenne):
    """Mention d'honneur du diplome, deduite de la moyenne annuelle selon le
    bareme usuel (Tunisie/LMD). En dessous de 10, aucune mention."""
    m = to_number(moyenne)
    if m >= 16:
        return "Très Bien"
    if m >= 14:
        return "Bien"
    if m >= 12:
        return "Assez Bien"
    if m >= 10:
        return "Passable"
    return ""


# ---------------------------------------------------------------------------
# Bloc UE/ECUE dynamique : le gabarit fournit deux lignes-modeles (1ere ligne
# d'une UE, et ligne de continuation pour une UE a plusieurs ECUE) que l'on
# duplique/style autant de fois que necessaire pour coller exactement a la
# structure UE/ECUE trouvee dans le PV, quelle que soit la filiere.
# ---------------------------------------------------------------------------

_STYLE_COLUMNS = range(1, 21)  # A..T


def _row_style_snapshot(ws, row):
    snapshot = {}
    for col in _STYLE_COLUMNS:
        cell = ws.cell(row=row, column=col)
        snapshot[col] = {
            "font": copy.copy(cell.font),
            "border": copy.copy(cell.border),
            "fill": copy.copy(cell.fill),
            "alignment": copy.copy(cell.alignment),
            "number_format": cell.number_format,
        }
    return snapshot


def _apply_row_style(ws, row, snapshot):
    for col, style in snapshot.items():
        cell = ws.cell(row=row, column=col)
        cell.font = style["font"]
        cell.border = style["border"]
        cell.fill = style["fill"]
        cell.alignment = style["alignment"]
        cell.number_format = style["number_format"]


def build_row_layout(ues_ordered):
    """Calcule, pour une liste d'UE (deja dans l'ordre du/des PV, donc
    regroupees par semestre), la ligne de depart de chaque UE dans le bloc
    dynamique, le nombre total de lignes ECUE, et les plages de lignes par
    semestre (pour la colonne "Sem")."""
    row_of = {}
    row = UE_BLOCK_FIRST_ROW
    spans = []
    if ues_ordered:
        seg_label = ues_ordered[0].semestre
        seg_start = row
        for ue in ues_ordered:
            if ue.semestre != seg_label:
                spans.append((seg_label, seg_start, row - 1))
                seg_label, seg_start = ue.semestre, row
            row_of[ue.label] = row
            row += len(ue.ecues)
        spans.append((seg_label, seg_start, row - 1))
    n_rows = row - UE_BLOCK_FIRST_ROW
    return row_of, n_rows, spans


def _write_dynamic_ue_block(ws, ues_ordered, row_of, n_rows, semester_spans):
    """Redimensionne le bloc UE/ECUE du gabarit (deja charge dans ws) pour
    contenir exactement n_rows lignes ECUE, puis ecrit le contenu structurel
    (libelles UE/ECUE, coefficients, credits) et les fusions correspondantes.
    Renvoie le decalage (diff) applique au pied de page."""
    style_first = _row_style_snapshot(ws, UE_BLOCK_FIRST_ROW)
    style_continuation = _row_style_snapshot(ws, UE_BLOCK_FIRST_ROW + 1)

    # openpyxl ne decale PAS les fusions existantes lors de insert_rows /
    # delete_rows (seules les valeurs/styles des cellules le sont) : on
    # retire nous-memes toutes les fusions a partir de la ligne 19 (bloc UE
    # d'origine + pied de page + la bordure decorative colonne O qui
    # traverse les deux), et on recree celles du pied de page / colonne O au
    # bon endroit une fois le bloc redimensionne. Celles du bloc UE
    # d'origine (19 a 39) sont abandonnees : on les reconstruit nous-memes
    # plus bas, adaptees a la structure UE/ECUE reelle.
    carried_merges = []  # (min_row, max_row, min_col, max_col, shift_min_row)
    for mc in list(ws.merged_cells.ranges):
        if mc.max_row < UE_BLOCK_FIRST_ROW:
            continue  # zone d'en-tete, jamais touchee
        ws.unmerge_cells(str(mc))
        if mc.min_row < UE_BLOCK_FIRST_ROW:
            # fusion qui traverse le bloc depuis l'en-tete (bordure colonne O) :
            # le haut ne bouge pas, seul le bas suit le bloc UE.
            carried_merges.append((mc.min_row, mc.max_row, mc.min_col, mc.max_col, False))
        elif mc.min_row >= FOOTER_FIRST_ROW:
            # fusion du pied de page : suit integralement le decalage.
            carried_merges.append((mc.min_row, mc.max_row, mc.min_col, mc.max_col, True))
        # sinon (19 <= min_row < 40) : ancienne fusion du bloc UE, abandonnee.

    diff = n_rows - (FOOTER_FIRST_ROW - UE_BLOCK_FIRST_ROW)
    if diff > 0:
        ws.insert_rows(FOOTER_FIRST_ROW, diff)
    elif diff < 0:
        ws.delete_rows(FOOTER_FIRST_ROW + diff, -diff)

    for orig_min_row, orig_max_row, min_col, max_col, shift_min_row in carried_merges:
        new_min_row = orig_min_row + diff if shift_min_row else orig_min_row
        ws.merge_cells(start_row=new_min_row, start_column=min_col,
                        end_row=orig_max_row + diff, end_column=max_col)

    for ue in ues_ordered:
        start = row_of[ue.label]
        n = len(ue.ecues)
        for offset, ec in enumerate(ue.ecues):
            r = start + offset
            _apply_row_style(ws, r, style_first if offset == 0 else style_continuation)
            ws.cell(row=r, column=3, value=ec.name)   # C : nom ECUE
            ws.cell(row=r, column=6, value=ec.coef)    # F : coef ECUE
            ws.cell(row=r, column=7, value=ec.credit)  # G : credit ECUE
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)  # C:D
        ws.cell(row=start, column=2, value=ue.label)   # B : nom UE
        ws.cell(row=start, column=8, value=ue.credit)  # H : credit UE
        if n > 1:
            ws.merge_cells(start_row=start, start_column=2, end_row=start + n - 1, end_column=2)
            ws.merge_cells(start_row=start, start_column=8, end_row=start + n - 1, end_column=8)
            for col in (17, 18, 19, 20):  # Q,R,S,T : synthese UE
                ws.merge_cells(start_row=start, start_column=col, end_row=start + n - 1, end_column=col)

    for label, start, end in semester_spans:
        digit = semester_digit(label)
        ws.cell(row=start, column=5, value=f"S{digit}" if digit else label)  # E : Sem
        if end > start:
            ws.merge_cells(start_row=start, start_column=5, end_row=end, end_column=5)

    return diff


# ---------------------------------------------------------------------------
# Fichier de coordonnees des etudiants (date/lieu de naissance, CIN, etc.)
# ---------------------------------------------------------------------------

def _normalize_sheet_key(name):
    return re.sub(r"[^A-Z0-9]", "", (name or "").upper())


def find_coordinates_sheet(wb, niveau, code):
    """Cherche l'onglet correspondant au niveau+filiere (ex. niveau='L1',
    code='BD' -> onglet 'L1-BD'). Renvoie None si aucun onglet ne correspond."""
    target = _normalize_sheet_key(f"{niveau}{code}")
    if not target:
        return None
    for name in wb.sheetnames:
        if target in _normalize_sheet_key(name):
            return wb[name]
    return None


def parse_student_coordinates(path, niveau, code):
    """Lit l'onglet niveau+filiere du fichier de coordonnees des etudiants.
    Renvoie un dict {nom_normalise: {naissance_date, naissance_lieu,
    nationalite, cin}}, ou {} si l'onglet est introuvable."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = find_coordinates_sheet(wb, niveau, code)
    if ws is None:
        return {}

    header_row = None
    for row in range(1, 20):
        val = ws.cell(row=row, column=2).value
        if val and "اسم" in str(val):
            header_row = row
            break
    if header_row is None:
        return {}

    coords = {}
    row = header_row + 2  # +1 = sous-entetes (Baccalaureat), +1 = 1ere ligne de donnees
    while True:
        prenom = ws.cell(row=row, column=2).value
        nom = ws.cell(row=row, column=3).value
        if not _clean(prenom) and not _clean(nom):
            break
        key = normalize_name(f"{_clean(nom)} {_clean(prenom)}")
        cin = ws.cell(row=row, column=8).value
        coords[key] = {
            "sexe": _clean(ws.cell(row=row, column=4).value),  # colonne "الجنس" : Femme / Homme
            "naissance_date": ws.cell(row=row, column=5).value,
            "naissance_lieu": _clean(ws.cell(row=row, column=6).value),
            "nationalite": _clean(ws.cell(row=row, column=7).value),
            "cin": _clean(cin) if cin is not None else "",
        }
        row += 1
    return coords


def civilite_from_sexe(sexe):
    """(civilite, prefixe_ne) selon le sexe : ('Mme', 'Née') pour une femme,
    ('M.', 'Né') sinon (defaut masculin si l'info est absente)."""
    if normalize_name(sexe).startswith("F"):  # Femme / Féminin
        return "Mme", "Née"
    return "M.", "Né"


def format_naissance(info):
    date_val = info.get("naissance_date")
    lieu = info.get("naissance_lieu") or ""
    date_str = date_val.strftime("%d/%m/%Y") if hasattr(date_val, "strftime") else _clean(date_val)
    if date_str and lieu:
        return f"{date_str} à {lieu}"
    return date_str or lieu


def fill_releve(template_path, output_path, nom_prenom, grades, ue_results,
                 ues_ordered, row_of, n_rows, semester_spans,
                 moyenne_annuelle, total_credits, meta, coordonnees=None):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    coordonnees = coordonnees or {}
    ws["D9"] = nom_prenom
    ws["D10"] = format_naissance(coordonnees)
    ws["D11"] = coordonnees.get("cin", "")
    ws["L8"] = f"Année universitaire : {meta['annee_universitaire']}"
    ws["C13"] = meta["diplome"]
    ws["C14"] = meta["domaine"]
    ws["L14"] = f"Niveau d'études : {meta['niveau']}"
    ws["C15"] = meta["specialite"]

    diff = _write_dynamic_ue_block(ws, ues_ordered, row_of, n_rows, semester_spans)

    for ue in ues_ordered:
        ue_row = row_of[ue.label]
        if ue.label in ue_results:
            moyenne_ue, credits_cap = ue_results[ue.label]
            credits_cap = int(to_number(credits_cap))
            ws[f"Q{ue_row}"] = credits_cap
            ws[f"R{ue_row}"] = "V" if credits_cap > 0 else None
            ws[f"S{ue_row}"] = "C" if credits_cap > 0 else None
            ws[f"T{ue_row}"] = meta["annee_universitaire"]
            for offset, ec in enumerate(ue.ecues):
                r = ue_row + offset
                cc, ex, moy, cr = grades.get(ec.code, (None, None, None, None))
                ws[f"J{r}"] = format_note(cc) if cc is not None else None
                ws[f"K{r}"] = format_note(ex) if ex is not None else None
                ws[f"L{r}"] = format_note(moy) if moy is not None else None
        # sinon : UE non couverte par les PV fournis -> deja vide (lignes neuves)

    total_row = FOOTER_FIRST_ROW + diff
    moyenne_row, controle_row = 42 + diff, 43 + diff
    credits_row = 45 + diff
    mention_row = 48 + diff
    print_end_row = 50 + diff

    total_credits_possible = ws[f"G{total_row}"].value
    ws[f"G{moyenne_row}"] = moyenne_annuelle
    ws[f"G{controle_row}"] = "-"
    ws[f"F{credits_row}"] = f"{total_credits}/{total_credits_possible}"
    ws[f"C{mention_row}"] = ""  # decision/mention : remplissage manuel par le jury
    # "Fait a Tunis : <date>" (meme ligne que la Moyenne Annuelle, colonne N).
    # Rempli seulement si une date est fournie ; sinon on garde celle du gabarit.
    if meta.get("date_remplissage"):
        ws[f"N{moyenne_row}"] = f"Fait à Tunis : {meta['date_remplissage']}"

    # Mise en page prete a l'impression (1 page A4).
    ws.print_area = f"A1:T{print_end_row}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.title = sanitize_sheet_name(nom_prenom)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def generate_all(pv_paths, template_path, output_dir, *,
                  coords_path=None, filiere_code=None, ar_template_path=None, niveau=None,
                  annee_universitaire=None, date_remplissage=None, diplome_templates=None):
    """Pipeline complet : parse 1..N PV, fusionne par etudiant, ecrit un
    releve (et une attestation de reussite si ar_template_path est fourni)
    par etudiant dans output_dir. Renvoie (fichiers_generes, avertissements).

    niveau (ex. "L1", "M2") est fourni explicitement par l'appelant (menu
    deroulant de l'interface) ; a defaut, on retombe sur une detection
    depuis le nom du 1er fichier PV (utile pour la CLI).

    annee_universitaire (ex. "2025-2026") et date_remplissage (ex.
    "30/06/2025", le "Fait a Tunis : ...") sont saisis dans l'interface ; a
    defaut l'annee est deduite du nom du fichier et la date reste celle du
    gabarit."""
    parsed = [parse_pv(p) for p in pv_paths]
    ues_per_pv = [p[0] for p in parsed]
    students_per_pv = [p[1] for p in parsed]
    pv_meta = parsed[0][2]

    annee_detectee, niveau_detecte = parse_filename_meta(pv_paths[0])
    annee_universitaire = annee_universitaire or annee_detectee
    niveau = niveau or niveau_detecte
    cycle = "Licence" if niveau.startswith("L") else "Mastère" if niveau.startswith("M") else ""
    meta = {
        "annee_universitaire": annee_universitaire,
        "date_remplissage": date_remplissage or "",
        "niveau": niveau,
        "domaine": pv_meta["domaine"],
        "filiere": pv_meta.get("filiere", ""),
        "specialite": pv_meta["specialite"],
        "diplome": f"{cycle} en {pv_meta['specialite']}".strip(),
    }

    ues_ordered = [ue for ues in ues_per_pv for ue in ues]
    ue_by_label = {ue.label: ue for ue in ues_ordered}
    row_of, n_rows, semester_spans = build_row_layout(ues_ordered)
    merged = merge_students(students_per_pv)

    coords = {}
    warnings = []
    if coords_path and filiere_code:
        coords = parse_student_coordinates(coords_path, niveau, filiere_code)
        if not coords:
            warnings.append(
                f"Fichier de coordonnées : aucun onglet trouvé pour {niveau} {filiere_code}."
            )

    ar = None
    if ar_template_path:
        import generate_attestation as ar  # import tardif : optionnel (depend de python-docx)

    # Diplome national : uniquement en annee terminale de cycle (L3 / M2) et si
    # un modele correspondant au cycle est fourni. Import tardif (python-docx).
    diplome = None
    diplome_template = None
    if diplome_templates and niveau in FINAL_YEARS:
        diplome_template = diplome_templates.get(cycle)
        if diplome_template:
            import generate_diplome as diplome
        else:
            warnings.append(
                f"Diplôme : aucun modèle {cycle} fourni "
                f"(déposez un fichier 'Diplome {cycle}...docx' dans le dossier des modèles)."
            )

    output_dir.mkdir(parents=True, exist_ok=True)
    generated = []
    for nom_prenom, rec in merged.items():
        moyenne, total_credits = annual_aggregates(rec["ue_results"], ue_by_label)
        etu_coords = coords.get(normalize_name(nom_prenom)) if coords else None
        if coords and not etu_coords:
            warnings.append(f"{nom_prenom} : coordonnées non trouvées dans le fichier fourni")

        out_path = output_dir / f"Releve_{sanitize_filename(nom_prenom)}.xlsx"
        fill_releve(template_path, out_path, nom_prenom, rec["grades"], rec["ue_results"],
                    ues_ordered, row_of, n_rows, semester_spans,
                    moyenne, total_credits, meta, coordonnees=etu_coords)
        generated.append(out_path)

        if ar_template_path:
            ar_path = output_dir / f"AR_{sanitize_filename(nom_prenom)}.docx"
            niveau_num, niveau_suffixe = NIVEAU_TEXTE.get(niveau, ("", ""))
            ar.fill_attestation(
                ar_template_path, ar_path, nom_prenom,
                naissance_date=(etu_coords or {}).get("naissance_date"),
                cin=(etu_coords or {}).get("cin", ""),
                nationalite=(etu_coords or {}).get("nationalite") or "",
                cycle=cycle,
                niveau_num=niveau_num,
                niveau_suffixe=niveau_suffixe,
                specialite=meta["specialite"],
            )

        if diplome is not None:
            dip_path = output_dir / f"Diplome_{sanitize_filename(nom_prenom)}.docx"
            civilite, ne = civilite_from_sexe((etu_coords or {}).get("sexe", ""))
            diplome.fill_diplome(
                diplome_template, dip_path,
                nom_prenom=nom_prenom,
                civilite=civilite,
                ne=ne,
                naissance_date=(etu_coords or {}).get("naissance_date"),
                naissance_lieu=(etu_coords or {}).get("naissance_lieu", ""),
                cin=(etu_coords or {}).get("cin", ""),
                annee_universitaire=meta["annee_universitaire"],
                domaine=meta["domaine"],
                mention=meta["filiere"],
                specialite=meta["specialite"],
                honneur=mention_from_average(moyenne),
                date_diplome=meta["date_remplissage"],
            )
            if mention_from_average(moyenne) == "":
                warnings.append(
                    f"{nom_prenom} : moyenne annuelle < 10 — diplôme généré sans mention "
                    f"(à écarter si l'étudiant·e n'est pas admis·e)"
                )

        if len(pv_paths) > 1 and len(rec["present_in"]) < len(pv_paths):
            missing = [i + 1 for i in range(len(pv_paths)) if i not in rec["present_in"]]
            warnings.append(f"{nom_prenom} : absent du PV Semestre {', '.join(map(str, missing))}")
        if rec["grades"] and all(g[2] == 0 for g in rec["grades"].values()):
            warnings.append(f"{nom_prenom} : toutes les notes sont a 0 (probablement absent(e))")

    return generated, warnings


def main():
    import argparse
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("pv", nargs="*", help="PV de deliberation (S1, et optionnellement S2)")
    parser.add_argument("--template", type=Path, default=None)
    parser.add_argument("--out", type=Path, default=BASE_DIR / "releves_generes")
    args = parser.parse_args()

    template_path = args.template or next(BASE_DIR.glob("Exemple*.xlsx"))
    pv_paths = [Path(p) for p in args.pv] if args.pv else [next(BASE_DIR.glob("PV*.xlsx"))]

    generated, warnings = generate_all(pv_paths, template_path, args.out)

    print(f"{len(generated)} releve(s) genere(s) dans {args.out}")
    for path in generated:
        print(f"  - {path.name}")
    if warnings:
        print("\nA verifier :")
        for w in warnings:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
